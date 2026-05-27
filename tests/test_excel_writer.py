from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook, load_workbook

from hwpreader.excel_writer import get_processed_sources, save_outputs
from hwpreader.models import ExtractedRecord


class ExcelWriterTest(unittest.TestCase):
    def test_save_outputs_groups_records_by_week_in_one_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            first_file = root / "20260504.hwp"
            second_file = root / "20260511.hwp"
            first_file.write_text("first", encoding="utf-8")
            second_file.write_text("second", encoding="utf-8")
            output_path = root / "result" / "sorted_data.xlsx"

            records = [
                ExtractedRecord(
                    source_file=first_file,
                    values={"index": "1", "address": "Gyeongbuk Pohang"},
                ),
                ExtractedRecord(
                    source_file=second_file,
                    values={"index": "2", "address": "Gyeongnam Changwon"},
                ),
                ExtractedRecord(
                    source_file=second_file,
                    values={"index": "3", "address": "Seoul"},
                ),
            ]

            save_outputs(
                records=records,
                fields=["index", "address"],
                template_path=root / "missing_template.xlsx",
                output_path=output_path,
                updated_files=[first_file, second_file],
                address_prefixes=["Gyeong"],
                address_field="address",
            )

            workbook = load_workbook(output_path)
            try:
                self.assertIn("2026-05-04~05-08", workbook.sheetnames)
                self.assertIn("2026-05-11~05-15", workbook.sheetnames)
                self.assertNotIn(
                    "Seoul",
                    [cell.value for cell in workbook["2026-05-11~05-15"]["B"]],
                )
                self.assertEqual(
                    get_processed_sources(output_path),
                    {
                        first_file.name: f"{first_file.stat().st_size}:{first_file.stat().st_mtime_ns}",
                        second_file.name: f"{second_file.stat().st_size}:{second_file.stat().st_mtime_ns}",
                    },
                )
            finally:
                workbook.close()

    def test_save_outputs_replaces_changed_source_records(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source_file = root / "20260504.hwp"
            output_path = root / "result" / "sorted_data.xlsx"
            source_file.write_text("before", encoding="utf-8")

            save_outputs(
                records=[
                    ExtractedRecord(
                        source_file=source_file,
                        values={"index": "1", "address": "Gyeongbuk Gyeongju"},
                    )
                ],
                fields=["index", "address"],
                template_path=root / "missing_template.xlsx",
                output_path=output_path,
                updated_files=[source_file],
                address_prefixes=["Gyeong"],
                address_field="address",
            )

            source_file.write_text("after", encoding="utf-8")
            save_outputs(
                records=[
                    ExtractedRecord(
                        source_file=source_file,
                        values={"index": "2", "address": "Gyeongbuk Gumi"},
                    )
                ],
                fields=["index", "address"],
                template_path=root / "missing_template.xlsx",
                output_path=output_path,
                updated_files=[source_file],
                address_prefixes=["Gyeong"],
                address_field="address",
            )

            workbook = load_workbook(output_path)
            try:
                sheet = workbook["2026-05-04~05-08"]
                self.assertEqual(sheet["A2"].value, "2")
                self.assertEqual(sheet.max_row, 2)
            finally:
                workbook.close()

    def test_save_outputs_without_prefixes_keeps_all_records(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source_file = root / "20260504.hwp"
            source_file.write_text("source", encoding="utf-8")
            output_path = root / "result" / "full_data.xlsx"

            save_outputs(
                records=[
                    ExtractedRecord(
                        source_file=source_file,
                        values={"index": "1", "address": "Gyeongbuk Pohang"},
                    ),
                    ExtractedRecord(
                        source_file=source_file,
                        values={"index": "2", "address": "Seoul"},
                    ),
                ],
                fields=["index", "address"],
                template_path=root / "missing_template.xlsx",
                output_path=output_path,
                updated_files=[source_file],
            )

            workbook = load_workbook(output_path)
            try:
                sheet = workbook["2026-05-04~05-08"]
                self.assertEqual(sheet["A2"].value, "1")
                self.assertEqual(sheet["A3"].value, "2")
                self.assertEqual(sheet.max_row, 3)
            finally:
                workbook.close()

    def test_save_outputs_preserves_non_generated_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source_file = root / "20260511.hwp"
            source_file.write_text("source", encoding="utf-8")
            output_path = root / "result" / "sorted_data.xlsm"
            output_path.parent.mkdir(parents=True)

            workbook = Workbook()
            button_sheet = workbook.active
            button_sheet.title = "Buttons"
            button_sheet["A1"] = "Run macro"
            workbook.create_sheet("2026-05-04~05-08")
            workbook.save(output_path)
            workbook.close()

            save_outputs(
                records=[
                    ExtractedRecord(
                        source_file=source_file,
                        values={"index": "1", "address": "Gyeongbuk"},
                    )
                ],
                fields=["index", "address"],
                template_path=root / "missing_template.xlsx",
                output_path=output_path,
                updated_files=[source_file],
                address_prefixes=["Gyeong"],
                address_field="address",
            )

            workbook = load_workbook(output_path, keep_vba=True)
            try:
                self.assertEqual(workbook.sheetnames[0], "Buttons")
                self.assertEqual(workbook["Buttons"]["A1"].value, "Run macro")
                self.assertNotIn("2026-05-04~05-08", workbook.sheetnames)
                self.assertIn("2026-05-11~05-15", workbook.sheetnames)
            finally:
                workbook.close()

    def test_save_outputs_reuses_existing_week_sheet_widths(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source_file = root / "20260511.hwp"
            source_file.write_text("source", encoding="utf-8")
            output_path = root / "result" / "sorted_data.xlsx"
            output_path.parent.mkdir(parents=True)

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "2026-05-04~05-08"
            sheet.column_dimensions["A"].width = 5
            sheet.column_dimensions["B"].width = 31
            sheet.column_dimensions["I"].width = 18
            workbook.save(output_path)
            workbook.close()

            save_outputs(
                records=[
                    ExtractedRecord(
                        source_file=source_file,
                        values={"index": "1", "address": "Gyeongbuk"},
                    )
                ],
                fields=["index", "address"],
                template_path=root / "missing_template.xlsx",
                output_path=output_path,
                updated_files=[source_file],
            )

            workbook = load_workbook(output_path)
            try:
                sheet = workbook["2026-05-11~05-15"]
                self.assertEqual(sheet.column_dimensions["A"].width, 5)
                self.assertEqual(sheet.column_dimensions["B"].width, 31)
                self.assertEqual(sheet.column_dimensions["I"].width, 18)
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
