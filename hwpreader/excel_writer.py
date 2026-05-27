from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
import re
import shutil

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .models import ExtractedRecord


RECORDS_SHEET = "__records"
MANIFEST_SHEET = "__processed_sources"
EMPTY_SHEET = "정리본"
INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")
WEEK_SHEET_PATTERN = re.compile(r"^\d{4}-\d{2}-\d{2}~\d{2}-\d{2}$")


def get_processed_sources(output_path: Path) -> dict[str, str]:
    if not output_path.exists():
        return {}

    workbook = load_workbook(
        output_path,
        read_only=True,
        data_only=True,
        keep_vba=_keeps_vba(output_path),
    )
    try:
        if MANIFEST_SHEET not in workbook.sheetnames:
            return {}
        sheet = workbook[MANIFEST_SHEET]
        rows = sheet.iter_rows(min_row=2, values_only=True)
        return {
            str(source_file): str(signature)
            for source_file, signature in rows
            if source_file and signature
        }
    finally:
        workbook.close()


def load_records(output_path: Path, fields: list[str]) -> list[ExtractedRecord]:
    if not output_path.exists():
        return []

    workbook = load_workbook(
        output_path,
        read_only=True,
        data_only=True,
        keep_vba=_keeps_vba(output_path),
    )
    try:
        return _load_stored_records(workbook, fields)
    finally:
        workbook.close()


def save_outputs(
    records: list[ExtractedRecord],
    fields: list[str],
    template_path: Path,
    output_path: Path,
    updated_files: list[Path],
    address_prefixes: list[str] | None = None,
    replace_existing: bool = False,
    address_field: str = "주소",
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = _load_workbook(output_path, template_path)
    updated_names = {path.name for path in updated_files}

    kept_records = []
    if not replace_existing:
        kept_records = [
            record
            for record in _load_stored_records(workbook, fields)
            if record.source_file.name not in updated_names
        ]
    new_records = _filter_records(records, address_field, address_prefixes)
    stored_records = kept_records + new_records

    _rewrite_record_sheet(workbook, stored_records, fields)
    _rewrite_manifest_sheet(workbook, updated_files)
    _rewrite_weekly_sheets(workbook, stored_records, fields)
    workbook.save(output_path)
    return output_path


def _load_workbook(output_path: Path, template_path: Path) -> Workbook:
    if output_path.exists():
        return load_workbook(output_path, keep_vba=_keeps_vba(output_path))

    if template_path.exists():
        temp_path = template_path.parent / f".tmp_{template_path.name}"
        shutil.copy2(template_path, temp_path)
        try:
            return load_workbook(temp_path, keep_vba=_keeps_vba(template_path))
        finally:
            temp_path.unlink(missing_ok=True)

    return Workbook()


def _keeps_vba(path: Path) -> bool:
    return path.suffix.lower() == ".xlsm"


def _load_stored_records(workbook: Workbook, fields: list[str]) -> list[ExtractedRecord]:
    if RECORDS_SHEET not in workbook.sheetnames:
        return []

    sheet = workbook[RECORDS_SHEET]
    headers = [str(cell.value or "") for cell in sheet[1]]
    if "source_file" not in headers:
        return []

    source_index = headers.index("source_file")
    field_indexes = {
        field: headers.index(field)
        for field in fields
        if field in headers
    }
    records: list[ExtractedRecord] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if source_index >= len(row) or not row[source_index]:
            continue
        values = {
            field: str(row[index] or "")
            for field, index in field_indexes.items()
        }
        records.append(ExtractedRecord(source_file=Path(str(row[source_index])), values=values))
    return records


def _rewrite_record_sheet(
    workbook: Workbook,
    records: list[ExtractedRecord],
    fields: list[str],
) -> None:
    sheet = _get_or_create_sheet(workbook, RECORDS_SHEET)
    _clear_sheet(sheet)
    sheet.append(["source_file", *fields])
    for record in records:
        sheet.append([record.source_file.name, *[record.values.get(field, "") for field in fields]])
    sheet.sheet_state = "hidden"


def _rewrite_manifest_sheet(workbook: Workbook, updated_files: list[Path]) -> None:
    processed_sources = get_processed_sources_from_workbook(workbook)
    for path in updated_files:
        processed_sources[path.name] = _source_signature(path)

    sheet = _get_or_create_sheet(workbook, MANIFEST_SHEET)
    _clear_sheet(sheet)
    sheet.append(["source_file", "signature"])
    for source_file, signature in sorted(processed_sources.items()):
        sheet.append([source_file, signature])
    sheet.sheet_state = "hidden"


def get_processed_sources_from_workbook(workbook: Workbook) -> dict[str, str]:
    if MANIFEST_SHEET not in workbook.sheetnames:
        return {}
    sheet = workbook[MANIFEST_SHEET]
    return {
        str(source_file): str(signature)
        for source_file, signature in sheet.iter_rows(min_row=2, values_only=True)
        if source_file and signature
    }


def _rewrite_weekly_sheets(
    workbook: Workbook,
    records: list[ExtractedRecord],
    fields: list[str],
) -> None:
    widths = _generated_sheet_widths(workbook)
    _remove_generated_sheets(workbook)
    grouped = _group_records_by_week(records)

    if not grouped:
        sheet = workbook.create_sheet(EMPTY_SHEET)
        _rewrite_summary_sheet(sheet, [], fields, widths)
        return

    for week_start, week_records in sorted(grouped.items()):
        sheet = workbook.create_sheet(_week_sheet_name(week_start))
        _rewrite_summary_sheet(sheet, week_records, fields, widths)


def _generated_sheet_widths(workbook: Workbook) -> dict[str, float]:
    for sheet in workbook.worksheets:
        if WEEK_SHEET_PATTERN.match(sheet.title):
            return {
                column_letter: dimension.width
                for column_letter, dimension in sheet.column_dimensions.items()
                if dimension.width is not None
            }
    return {}


def _remove_generated_sheets(workbook: Workbook) -> None:
    for sheet in list(workbook.worksheets):
        if sheet.title == EMPTY_SHEET or WEEK_SHEET_PATTERN.match(sheet.title):
            workbook.remove(sheet)


def _group_records_by_week(
    records: list[ExtractedRecord],
) -> dict[datetime, list[ExtractedRecord]]:
    grouped: dict[datetime, list[ExtractedRecord]] = defaultdict(list)
    for record in records:
        source_date = _date_from_source_file(record.source_file)
        week_start = source_date - timedelta(days=source_date.weekday())
        grouped[week_start].append(record)
    return grouped


def _date_from_source_file(source_file: Path) -> datetime:
    return datetime.strptime(source_file.stem, "%Y%m%d")


def _week_sheet_name(week_start: datetime) -> str:
    week_end = week_start + timedelta(days=4)
    title = f"{week_start:%Y-%m-%d}~{week_end:%m-%d}"
    return INVALID_SHEET_CHARS.sub("-", title)[:31]


def _address_starts_with(
    record: ExtractedRecord,
    address_field: str,
    address_prefixes: list[str],
) -> bool:
    address = record.values.get(address_field, "").strip()
    return any(address.startswith(prefix) for prefix in address_prefixes)


def _filter_records(
    records: list[ExtractedRecord],
    address_field: str,
    address_prefixes: list[str] | None,
) -> list[ExtractedRecord]:
    if address_prefixes is None:
        return records
    return [
        record
        for record in records
        if _address_starts_with(record, address_field, address_prefixes)
    ]


def _rewrite_summary_sheet(
    sheet: Worksheet,
    records: list[ExtractedRecord],
    fields: list[str],
    widths: dict[str, float] | None = None,
) -> None:
    _clear_sheet(sheet)
    headers = fields + ["source_file"]
    sheet.append(headers)
    for record in records:
        sheet.append([record.values.get(field, "") for field in fields] + [record.source_file.name])
    _style_header(sheet)
    if widths:
        _apply_widths(sheet, widths)
    else:
        _set_widths(sheet, headers)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _clear_sheet(sheet: Worksheet) -> None:
    if sheet.max_row:
        sheet.delete_rows(1, sheet.max_row)


def _get_or_create_sheet(workbook: Workbook, name: str) -> Worksheet:
    if name in workbook.sheetnames:
        return workbook[name]
    return workbook.create_sheet(name)


def _style_header(sheet: Worksheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill


def _set_widths(sheet: Worksheet, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for row_idx in range(2, min(sheet.max_row, 80) + 1):
            value = sheet.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(value or "")))
        sheet.column_dimensions[sheet.cell(row=1, column=col_idx).column_letter].width = min(max(max_len + 2, 10), 42)


def _apply_widths(sheet: Worksheet, widths: dict[str, float]) -> None:
    for column_letter, width in widths.items():
        sheet.column_dimensions[column_letter].width = width


def _source_signature(path: Path) -> str:
    stat = path.stat()
    return f"{stat.st_size}:{stat.st_mtime_ns}"
